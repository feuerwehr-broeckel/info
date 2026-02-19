"""Import passive Fox112 members (Fördernde Mitglieder) into FeuerON.

Reads the Fox112 Excel export, filters to "Fördernde Mitglieder", and
generates a FeuerON CSV import file for persons not yet in FeuerON.

Persons with historical Dienstgrade or Funktionen (from the Fox112 XML
export) are skipped — they were formerly active and must be added manually.
"""

from __future__ import annotations

import logging
import re
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import NamedTuple

import openpyxl
import typer

from scripts.feueron.migration.contact_check import _format_phone
from scripts.feueron.migration.erreichbarkeiten_parser import parse_erreichbarkeiten
from scripts.feueron.models import (
    BeitragArt,
    Geschlecht,
    PersonRecord,
    PvAbt,
    PvBank,
    PvBeitrag,
    PvDb,
    PvTelep,
    TelArt,
    generate_csv,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Fox112 Excel reader
# ---------------------------------------------------------------------------


class _Fox112PassivePerson(NamedTuple):
    nachname: str
    vorname: str
    geschlecht: str
    geburtstag: str
    eintritt: str
    personal_nr: str
    strasse: str
    hausnr: str
    plz: str
    ort: str
    tel_p: str
    tel_d: str
    mobil_p: str
    mobil_d: str
    email_p: str
    email_d: str
    fax_p: str
    fax_d: str
    mitgliedsbeitrag: str
    zahlart: str
    kontoinhaber: str
    iban: str
    bic: str
    mandatsreferenz: str


def _read_fox112_excel(path: Path) -> list[_Fox112PassivePerson]:
    """Read passive members from a Fox112 Excel export."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers) if name}

    required = {
        "NACHNAME",
        "VORNAME",
        "GESCHLECHT",
        "GEBURTSTAG",
        "ABTEILUNG",
        "EINTRITT",
        "PERSONAL_NR",
        "STRASSE",
        "HAUSNR",
        "PLZ",
        "ORT",
        "TEL_P",
        "TEL_D",
        "MOBIL_P",
        "MOBIL_D",
        "EMAIL_P",
        "EMAIL_D",
        "FAX_P",
        "FAX_D",
        "MITGLIEDSBEITRAG",
        "ZAHLART",
        "KONTOINHABER",
        "IBAN",
        "BIC",
        "MANDATSREFERENZ",
    }
    missing = required - col.keys()
    if missing:
        msg = f"Fox112 Excel is missing columns: {', '.join(sorted(missing))}"
        raise ValueError(msg)

    persons: list[_Fox112PassivePerson] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nachname = str(row[col["NACHNAME"]] or "").strip()
        if not nachname:
            continue
        abteilung = str(row[col["ABTEILUNG"]] or "").strip()
        if abteilung != "Fördernde Mitglieder":
            continue
        persons.append(
            _Fox112PassivePerson(
                nachname=nachname,
                vorname=str(row[col["VORNAME"]] or "").strip(),
                geschlecht=str(row[col["GESCHLECHT"]] or "").strip(),
                geburtstag=str(row[col["GEBURTSTAG"]] or "").strip(),
                eintritt=str(row[col["EINTRITT"]] or "").strip(),
                personal_nr=str(row[col["PERSONAL_NR"]] or "").strip(),
                strasse=str(row[col["STRASSE"]] or "").strip(),
                hausnr=str(row[col["HAUSNR"]] or "").strip(),
                plz=str(row[col["PLZ"]] or "").strip(),
                ort=str(row[col["ORT"]] or "").strip(),
                tel_p=str(row[col["TEL_P"]] or "").strip(),
                tel_d=str(row[col["TEL_D"]] or "").strip(),
                mobil_p=str(row[col["MOBIL_P"]] or "").strip(),
                mobil_d=str(row[col["MOBIL_D"]] or "").strip(),
                email_p=str(row[col["EMAIL_P"]] or "").strip(),
                email_d=str(row[col["EMAIL_D"]] or "").strip(),
                fax_p=str(row[col["FAX_P"]] or "").strip(),
                fax_d=str(row[col["FAX_D"]] or "").strip(),
                mitgliedsbeitrag=str(row[col["MITGLIEDSBEITRAG"]] or "").strip(),
                zahlart=str(row[col["ZAHLART"]] or "").strip(),
                kontoinhaber=str(row[col["KONTOINHABER"]] or "").strip(),
                iban=str(row[col["IBAN"]] or "").strip(),
                bic=str(row[col["BIC"]] or "").strip(),
                mandatsreferenz=str(row[col["MANDATSREFERENZ"]] or "").strip(),
            )
        )

    wb.close()
    return persons


# ---------------------------------------------------------------------------
# Fox112 XML history reader
# ---------------------------------------------------------------------------

_NAME_PREFIX_RE = re.compile(r"^#\d*\s*")


def _read_xml_history(path: Path) -> set[tuple[str, str]]:
    """Read the Fox112 XML and return names of persons with Dienstgrade or Funktionen.

    Names are lowercased for matching.  The ``#`` / ``#N`` prefix in XML
    last names is stripped.
    """
    tree = ET.parse(path)  # noqa: S314
    root = tree.getroot()

    has_history: set[tuple[str, str]] = set()
    for person in root.iter("Person"):
        dienstgrade = person.find("Dienstgrade")
        funktionen = person.find("Funktionen")
        dg_entries = dienstgrade.findall("Eintrag") if dienstgrade is not None else []
        fn_entries = funktionen.findall("Eintrag") if funktionen is not None else []

        if not dg_entries and not fn_entries:
            continue

        nachname_el = person.find("Nachname")
        vorname_el = person.find("Vorname")
        if nachname_el is None or vorname_el is None:
            continue

        nachname = _NAME_PREFIX_RE.sub("", (nachname_el.text or "").strip()).lower()
        vorname = (vorname_el.text or "").strip().lower()
        has_history.add((nachname, vorname))

    return has_history


# ---------------------------------------------------------------------------
# Beitrag helpers
# ---------------------------------------------------------------------------

_ZAHLART_MAP = {
    "Einzug": BeitragArt.LASTSCHRIFT,
    "Barzahler": BeitragArt.KASSIERER,
    "Selbstzahler": BeitragArt.SELBSTUEBERWEISER,
}


def _parse_beitrag(raw: str) -> float:
    """Parse a Fox112 MITGLIEDSBEITRAG value like ``"20,00 €"`` to float."""
    cleaned = raw.replace("€", "").replace(",", ".").strip()
    if not cleaned:
        return 0.0
    return float(cleaned)


# ---------------------------------------------------------------------------
# Contact field mapping (same as contact_check.py)
# ---------------------------------------------------------------------------

_CONTACT_FIELDS: list[tuple[str, TelArt, bool]] = [
    ("tel_p", TelArt.TELEFON_PRIVAT, True),
    ("tel_d", TelArt.TELEFON_DIENSTLICH, True),
    ("mobil_p", TelArt.MOBIL_PRIVAT, True),
    ("mobil_d", TelArt.MOBIL_DIENSTLICH, True),
    ("email_p", TelArt.EMAIL_PRIVAT, False),
    ("email_d", TelArt.EMAIL_DIENSTLICH, False),
    ("fax_p", TelArt.TELEFAX_PRIVAT, True),
    ("fax_d", TelArt.TELEFAX_DIENSTLICH, True),
]


# ---------------------------------------------------------------------------
# Core: build passive-import CSV
# ---------------------------------------------------------------------------

_GESCHLECHT_MAP = {
    "M": Geschlecht.MAENNLICH,
    "W": Geschlecht.WEIBLICH,
    "J": Geschlecht.JURISTISCH,
}


def build_passive_import(
    fox112_excel: Path,
    feueron_erreichbarkeiten: Path,
    fox112_xml: Path,
    output: Path,
    output_no_beitrag: Path,
    *,
    organisation: str = "Bröckel, OF",
    yes: bool = False,
) -> None:
    """Build FeuerON CSVs to import passive members from Fox112.

    Reads "Fördernde Mitglieder" from the Fox112 Excel, skips persons
    already in FeuerON (by name match against Erreichbarkeiten report)
    and persons with historical Dienstgrade/Funktionen (from the Fox112
    XML export).

    Produces two CSV files: ``output`` for persons with Beiträge data, and
    ``output_no_beitrag`` for persons without (they must be imported
    separately because FeuerON rejects rows with empty PV_BEITRAG columns).
    """
    # 1. Parse inputs
    fox112_persons = _read_fox112_excel(fox112_excel)
    logger.info("Fox112: %d Fördernde Mitglieder", len(fox112_persons))

    feueron_entries = parse_erreichbarkeiten(feueron_erreichbarkeiten)
    feueron_names: set[tuple[str, str]] = {
        (e.nachname.lower(), e.vorname.lower()) for e in feueron_entries
    }
    logger.info("FeuerON: %d persons already present", len(feueron_names))

    xml_history = _read_xml_history(fox112_xml)
    logger.info("XML: %d persons with Dienstgrade/Funktionen", len(xml_history))

    # 2. Determine next PERSONAL_NR for persons that lack one
    _PERS_NR_RE = re.compile(r"^Broe_F_(\d+)$")
    max_nr = 0
    for fox in fox112_persons:
        m = _PERS_NR_RE.match(fox.personal_nr)
        if m:
            max_nr = max(max_nr, int(m.group(1)))
    next_nr = max_nr + 1

    # 3. Build PersonRecords
    records: list[PersonRecord] = []
    included = 0
    skipped_in_feueron = 0
    skipped_has_history = 0
    skipped_declined = 0
    skipped_no_geschlecht = 0
    skipped_missing_data = 0

    for fox in fox112_persons:
        key = (fox.nachname.lower(), fox.vorname.lower())

        if key in feueron_names:
            skipped_in_feueron += 1
            continue

        if key in xml_history:
            logger.warning(
                "%s, %s — has Dienstgrade/Funktionen, add manually",
                fox.nachname,
                fox.vorname,
            )
            skipped_has_history += 1
            continue

        geschlecht = _GESCHLECHT_MAP.get(fox.geschlecht.upper())
        if geschlecht is None:
            logger.warning(
                "Unknown GESCHLECHT '%s' for %s, %s — skipped",
                fox.geschlecht,
                fox.nachname,
                fox.vorname,
            )
            skipped_no_geschlecht += 1
            continue

        if not fox.geburtstag or not fox.eintritt:
            logger.warning(
                "Missing GEBURTSTAG or EINTRITT for %s, %s — skipped",
                fox.nachname,
                fox.vorname,
            )
            skipped_missing_data += 1
            continue

        # Build contact entries
        erreichbarkeiten: list[PvTelep] = []
        for fox_field, tel_art, is_phone in _CONTACT_FIELDS:
            value = getattr(fox, fox_field)
            if value:
                formatted = _format_phone(value) if is_phone else value
                erreichbarkeiten.append(PvTelep(TEL_ART=tel_art, TELEPHON=formatted))

        # Build bank entry
        bank: PvBank | None = None
        if fox.iban or fox.bic or fox.kontoinhaber or fox.mandatsreferenz:
            bank = PvBank(
                INHABER=fox.kontoinhaber or None,
                IBAN=fox.iban or None,
                BIC=fox.bic or None,
                MANDATSREFERENZ=fox.mandatsreferenz or None,
            )

        # Build Beitrag entry
        beitraege: list[PvBeitrag] = []
        betrag = _parse_beitrag(fox.mitgliedsbeitrag)
        beitrag_art = _ZAHLART_MAP.get(fox.zahlart)
        if betrag > 0 and beitrag_art is not None and fox.eintritt:
            beitraege.append(
                PvBeitrag(
                    BETRAG=betrag,
                    TYP="Mitgliedsbeitrag",
                    ART=beitrag_art,
                    GULTIG_AB=fox.eintritt,
                    ERSTE_FAELLIGKEIT=fox.eintritt,
                )
            )

        # Resolve PERS_NR
        pers_nr = fox.personal_nr
        if not pers_nr:
            pers_nr = f"Broe_F_{next_nr:03d}"
            next_nr += 1

        # Log and confirm
        logger.info(
            "Importing %s, %s (%s, geb. %s, Eintritt %s)",
            fox.nachname,
            fox.vorname,
            pers_nr,
            fox.geburtstag,
            fox.eintritt,
        )

        if not yes and not typer.confirm("  Include in CSV?", default=True):
            skipped_declined += 1
            continue

        records.append(
            PersonRecord(
                stammdaten=PvDb(
                    ORGANISATION=organisation,
                    NACHNAME=fox.nachname,
                    VORNAME=fox.vorname,
                    PERS_NR=pers_nr,
                    GEBURT=fox.geburtstag,
                    GESCHLECHT=geschlecht,
                    STRASSE=fox.strasse or None,
                    HAUSNR=fox.hausnr or None,
                    PLZ=fox.plz or None,
                    ORT=fox.ort or None,
                ),
                abteilungen=[
                    PvAbt(
                        ABTEILUNG="Fördernde Mitglieder",
                        VON=fox.eintritt,
                        BUNDESLAND="Niedersachsen",
                    ),
                ],
                erreichbarkeiten=erreichbarkeiten,
                bank=bank,
                beitraege=beitraege,
            )
        )
        included += 1

    # 4. Split into records with and without Beiträge
    records_with_beitrag = [r for r in records if r.beitraege]
    records_no_beitrag = [r for r in records if not r.beitraege]

    logger.info(
        "Passive import: %d included (%d with Beitrag, %d without), "
        "%d declined, %d already in FeuerON, "
        "%d have history (manual), %d missing data, %d skipped (no GESCHLECHT)",
        included,
        len(records_with_beitrag),
        len(records_no_beitrag),
        skipped_declined,
        skipped_in_feueron,
        skipped_has_history,
        skipped_missing_data,
        skipped_no_geschlecht,
    )

    # 5. Generate FeuerON CSVs
    generate_csv(records_with_beitrag, output)
    generate_csv(records_no_beitrag, output_no_beitrag)
