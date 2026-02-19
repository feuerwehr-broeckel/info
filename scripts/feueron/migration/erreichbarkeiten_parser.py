"""Parser for the FeuerON Erreichbarkeiten Excel report.

The Erreichbarkeiten report is only available as Excel (xlsx).  It has a flat
structure: metadata in rows 1–10, a header row at row 11, and one data row per
person from row 12 onward.
"""

from __future__ import annotations

from pathlib import Path
from typing import NamedTuple

import openpyxl


class ErreichbarkeitenEntry(NamedTuple):
    """A single person from a FeuerON Erreichbarkeiten report."""

    nachname: str
    vorname: str
    personal_nr: str
    geburtsdatum: str
    strasse: str
    hausnr: str
    plz: str
    ort: str
    ortsteil: str
    telefon_privat: str
    telefon_dienstlich: str
    telefax_privat: str
    telefax_dienstlich: str
    mobil_privat: str
    mobil_dienstlich: str
    email_privat: str
    email_dienstlich: str


# Column name → header row index (row 11 / 0-indexed row 10)
_COLUMNS = {
    "Name": "nachname",
    "Vorname": "vorname",
    "Personalnummer": "personal_nr",
    "Geburtsdatum": "geburtsdatum",
    "Straße": "strasse",
    "Hausnr.": "hausnr",
    "PLZ": "plz",
    "Ort": "ort",
    "Ortsteil": "ortsteil",
    "Telefon privat": "telefon_privat",
    "Telefon dienstlich": "telefon_dienstlich",
    "Telefax privat": "telefax_privat",
    "Telefax dienstlich": "telefax_dienstlich",
    "Mobil privat": "mobil_privat",
    "Mobil dienstlich": "mobil_dienstlich",
    "E-Mail privat": "email_privat",
    "E-Mail dienstlich": "email_dienstlich",
}


def parse_erreichbarkeiten(path: Path) -> list[ErreichbarkeitenEntry]:
    """Parse a FeuerON Erreichbarkeiten Excel report.

    Expects the header row at row 11 (1-indexed) and data from row 12 onward.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Read header at row 11
    headers = [cell.value for cell in next(ws.iter_rows(min_row=11, max_row=11))]
    col: dict[str, int] = {}
    for idx, h in enumerate(headers):
        if h and h in _COLUMNS:
            col[_COLUMNS[h]] = idx

    missing = set(_COLUMNS.values()) - col.keys()
    if missing:
        wb.close()
        msg = (
            f"Erreichbarkeiten report is missing columns: {', '.join(sorted(missing))}"
        )
        raise ValueError(msg)

    entries: list[ErreichbarkeitenEntry] = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        nachname = str(row[col["nachname"]] or "").strip()
        if not nachname:
            continue
        entries.append(
            ErreichbarkeitenEntry(
                nachname=nachname,
                vorname=str(row[col["vorname"]] or "").strip(),
                personal_nr=str(row[col["personal_nr"]] or "").strip(),
                geburtsdatum=str(row[col["geburtsdatum"]] or "").strip(),
                strasse=str(row[col["strasse"]] or "").strip(),
                hausnr=str(row[col["hausnr"]] or "").strip(),
                plz=str(row[col["plz"]] or "").strip(),
                ort=str(row[col["ort"]] or "").strip(),
                ortsteil=str(row[col["ortsteil"]] or "").strip(),
                telefon_privat=str(row[col["telefon_privat"]] or "").strip(),
                telefon_dienstlich=str(row[col["telefon_dienstlich"]] or "").strip(),
                telefax_privat=str(row[col["telefax_privat"]] or "").strip(),
                telefax_dienstlich=str(row[col["telefax_dienstlich"]] or "").strip(),
                mobil_privat=str(row[col["mobil_privat"]] or "").strip(),
                mobil_dienstlich=str(row[col["mobil_dienstlich"]] or "").strip(),
                email_privat=str(row[col["email_privat"]] or "").strip(),
                email_dienstlich=str(row[col["email_dienstlich"]] or "").strip(),
            )
        )

    wb.close()
    return entries
