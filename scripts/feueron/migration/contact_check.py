"""Build a FeuerON contact-update CSV from Fox112 master data.

Compares Fox112 addresses and contact details (Excel export) with the FeuerON
Erreichbarkeiten report (Excel) and outputs a FeuerON-compatible CSV import
file.  Fox112 is the source of truth; the FeuerON report provides the PERS_NR
and current values for diff logging.
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import NamedTuple

import openpyxl
import phonenumbers
import typer

from scripts.feueron.migration.erreichbarkeiten_parser import parse_erreichbarkeiten
from scripts.feueron.models import (
    Geschlecht,
    PersonRecord,
    PvDb,
    PvTelep,
    TelArt,
    generate_csv,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Fox112 Excel reader
# ---------------------------------------------------------------------------


class _Fox112ContactPerson(NamedTuple):
    nachname: str
    vorname: str
    geschlecht: str
    geburtstag: str
    strasse: str
    hausnr: str
    plz: str
    ort: str
    tel_p: str
    tel_d: str
    mobil_p: str
    mobil_d: str
    email_p: str
    email_d: str
    fax_p: str
    fax_d: str


def _read_fox112_excel(path: Path) -> list[_Fox112ContactPerson]:
    """Read address and contact columns from a Fox112 Excel export."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers) if name}

    required = {
        "NACHNAME",
        "VORNAME",
        "GESCHLECHT",
        "GEBURTSTAG",
        "ABTEILUNG",
        "STRASSE",
        "HAUSNR",
        "PLZ",
        "ORT",
        "TEL_P",
        "TEL_D",
        "MOBIL_P",
        "MOBIL_D",
        "EMAIL_P",
        "EMAIL_D",
        "FAX_P",
        "FAX_D",
    }
    missing = required - col.keys()
    if missing:
        msg = f"Fox112 Excel is missing columns: {', '.join(sorted(missing))}"
        raise ValueError(msg)

    persons: list[_Fox112ContactPerson] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nachname = str(row[col["NACHNAME"]] or "").strip()
        vorname = str(row[col["VORNAME"]] or "").strip()
        if not nachname:
            continue
        abteilung = str(row[col["ABTEILUNG"]] or "").strip()
        if abteilung == "Externe Kontakte":
            continue
        persons.append(
            _Fox112ContactPerson(
                nachname=nachname,
                vorname=vorname,
                geschlecht=str(row[col["GESCHLECHT"]] or "").strip(),
                geburtstag=str(row[col["GEBURTSTAG"]] or "").strip(),
                strasse=str(row[col["STRASSE"]] or "").strip(),
                hausnr=str(row[col["HAUSNR"]] or "").strip(),
                plz=str(row[col["PLZ"]] or "").strip(),
                ort=str(row[col["ORT"]] or "").strip(),
                tel_p=str(row[col["TEL_P"]] or "").strip(),
                tel_d=str(row[col["TEL_D"]] or "").strip(),
                mobil_p=str(row[col["MOBIL_P"]] or "").strip(),
                mobil_d=str(row[col["MOBIL_D"]] or "").strip(),
                email_p=str(row[col["EMAIL_P"]] or "").strip(),
                email_d=str(row[col["EMAIL_D"]] or "").strip(),
                fax_p=str(row[col["FAX_P"]] or "").strip(),
                fax_d=str(row[col["FAX_D"]] or "").strip(),
            )
        )

    wb.close()
    return persons


# ---------------------------------------------------------------------------
# Phone number formatting
# ---------------------------------------------------------------------------


def _format_phone(raw: str) -> str:
    """Format a phone number to national German format using phonenumbers.

    Returns the original string unchanged if parsing fails.
    """
    try:
        parsed = phonenumbers.parse(raw, "DE")
        return phonenumbers.format_number(
            parsed, phonenumbers.PhoneNumberFormat.NATIONAL
        )
    except phonenumbers.NumberParseException:
        return raw


# ---------------------------------------------------------------------------
# Address field mapping
# ---------------------------------------------------------------------------

# (Fox112 field, label, FeuerON Erreichbarkeiten field)
_ADDRESS_FIELDS: list[tuple[str, str, str]] = [
    ("strasse", "Straße", "strasse"),
    ("hausnr", "Hausnr.", "hausnr"),
    ("plz", "PLZ", "plz"),
    ("ort", "Ort", "ort"),
]


# ---------------------------------------------------------------------------
# Contact field mapping
# ---------------------------------------------------------------------------

# (Fox112 field, TelArt, FeuerON Erreichbarkeiten field, is_phone)
_CONTACT_FIELDS: list[tuple[str, TelArt, str, bool]] = [
    ("tel_p", TelArt.TELEFON_PRIVAT, "telefon_privat", True),
    ("tel_d", TelArt.TELEFON_DIENSTLICH, "telefon_dienstlich", True),
    ("mobil_p", TelArt.MOBIL_PRIVAT, "mobil_privat", True),
    ("mobil_d", TelArt.MOBIL_DIENSTLICH, "mobil_dienstlich", True),
    ("email_p", TelArt.EMAIL_PRIVAT, "email_privat", False),
    ("email_d", TelArt.EMAIL_DIENSTLICH, "email_dienstlich", False),
    ("fax_p", TelArt.TELEFAX_PRIVAT, "telefax_privat", True),
    ("fax_d", TelArt.TELEFAX_DIENSTLICH, "telefax_dienstlich", True),
]


# ---------------------------------------------------------------------------
# Core: build contact-update CSV
# ---------------------------------------------------------------------------

_GESCHLECHT_MAP = {
    "M": Geschlecht.MAENNLICH,
    "W": Geschlecht.WEIBLICH,
    "J": Geschlecht.JURISTISCH,
}


def build_contact_update(
    fox112_excel: Path,
    feueron_erreichbarkeiten: Path,
    output: Path | io.StringIO,
    *,
    organisation: str = "Bröckel, OF",
    yes: bool = False,
) -> None:
    """Build a FeuerON-compatible CSV to update addresses and contacts.

    Matches persons by name (case-insensitive).  Fox112 is the master for
    address and contact data; the FeuerON Erreichbarkeiten report provides
    the ``PERS_NR`` and current values for diff logging.
    """
    # 1. Parse inputs
    feueron_entries = parse_erreichbarkeiten(feueron_erreichbarkeiten)
    fox112_persons = _read_fox112_excel(fox112_excel)

    # 2. Build Fox112 name index
    fox_by_name: dict[tuple[str, str], list[_Fox112ContactPerson]] = {}
    for person in fox112_persons:
        key = (person.nachname.lower(), person.vorname.lower())
        fox_by_name.setdefault(key, []).append(person)

    # 3. Match and build PersonRecords
    records: list[PersonRecord] = []
    matched = 0
    skipped_no_match = 0
    skipped_duplicate = 0
    skipped_declined = 0
    skipped_unchanged = 0

    for entry in feueron_entries:
        key = (entry.nachname.lower(), entry.vorname.lower())
        candidates = fox_by_name.get(key, [])

        if len(candidates) == 0:
            logger.warning(
                "No Fox112 match for %s, %s (FeuerON PNR: %s) — skipped",
                entry.nachname,
                entry.vorname,
                entry.personal_nr,
            )
            skipped_no_match += 1
            continue

        if len(candidates) > 1:
            logger.warning(
                "Multiple Fox112 matches (%d) for %s, %s (FeuerON PNR: %s) — skipped",
                len(candidates),
                entry.nachname,
                entry.vorname,
                entry.personal_nr,
            )
            skipped_duplicate += 1
            continue

        fox = candidates[0]
        geschlecht = _GESCHLECHT_MAP.get(fox.geschlecht.upper())
        if geschlecht is None:
            logger.warning(
                "Unknown GESCHLECHT '%s' for %s, %s — skipped",
                fox.geschlecht,
                entry.nachname,
                entry.vorname,
            )
            continue

        # Detect stammdaten changes (birthday + address, per field)
        stammdaten_changes: list[tuple[str, str, str]] = []
        if fox.geburtstag != entry.geburtsdatum:
            stammdaten_changes.append(
                ("Geburtsdatum", entry.geburtsdatum, fox.geburtstag)
            )
        for fox_field, label, feueron_field in _ADDRESS_FIELDS:
            fox_val = getattr(fox, fox_field)
            feueron_val = getattr(entry, feueron_field)
            if fox_val != feueron_val:
                stammdaten_changes.append((label, feueron_val, fox_val))

        # Build contact entries (with phone formatting)
        erreichbarkeiten: list[PvTelep] = []
        for fox_field, tel_art, _, is_phone in _CONTACT_FIELDS:
            value = getattr(fox, fox_field)
            if value:
                formatted = _format_phone(value) if is_phone else value
                erreichbarkeiten.append(PvTelep(TEL_ART=tel_art, TELEPHON=formatted))

        # Detect contact changes (comparing formatted Fox112 values to FeuerON)
        contact_changes: list[tuple[str, str, str]] = []
        for fox_field, tel_art, feueron_field, is_phone in _CONTACT_FIELDS:
            fox_val = getattr(fox, fox_field)
            if not fox_val:
                continue
            formatted = _format_phone(fox_val) if is_phone else fox_val
            feueron_val = getattr(entry, feueron_field)
            if formatted != feueron_val:
                contact_changes.append((tel_art.value, feueron_val, formatted))

        # Skip persons with no changes
        if not stammdaten_changes and not contact_changes:
            skipped_unchanged += 1
            continue

        # Log and confirm changes
        logger.info(
            "Updating %s, %s (%s):",
            fox.nachname,
            fox.vorname,
            entry.personal_nr,
        )
        for label, old, new in stammdaten_changes:
            if old:
                logger.info("  %s: %s → %s", label, old, new)
            else:
                logger.info("  %s: (empty) → %s", label, new)
        for label, old, new in contact_changes:
            if old:
                logger.info("  %s: %s → %s", label, old, new)
            else:
                logger.info("  %s: (empty) → %s", label, new)

        if not yes and not typer.confirm("  Include in CSV?", default=True):
            skipped_declined += 1
            continue

        records.append(
            PersonRecord(
                stammdaten=PvDb(
                    ORGANISATION=organisation,
                    NACHNAME=fox.nachname,
                    VORNAME=fox.vorname,
                    PERS_NR=entry.personal_nr,
                    GEBURT=fox.geburtstag,
                    GESCHLECHT=geschlecht,
                    STRASSE=fox.strasse or None,
                    HAUSNR=fox.hausnr or None,
                    PLZ=fox.plz or None,
                    ORT=fox.ort or None,
                    ORTSTEIL=entry.ortsteil or None,
                ),
                erreichbarkeiten=erreichbarkeiten,
            )
        )
        matched += 1

    logger.info(
        "Contact update: %d included, %d declined, %d unchanged, %d skipped (no match), %d skipped (duplicate)",
        matched,
        skipped_declined,
        skipped_unchanged,
        skipped_no_match,
        skipped_duplicate,
    )

    # 4. Generate FeuerON CSV
    generate_csv(records, output)
