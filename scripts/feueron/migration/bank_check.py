"""Build a FeuerON bank-detail-update CSV from Fox112 master data.

Compares Fox112 bank details (Excel export) with the FeuerON
Bankverbindungen report and outputs a FeuerON-compatible CSV import file
to correct bank details where they differ.  Fox112 is the source of truth;
the FeuerON report provides the PERS_NR required for the import.
"""

from __future__ import annotations

import io
import logging
from pathlib import Path
from typing import NamedTuple

import openpyxl
import typer

from scripts.feueron.migration.report_parser import iter_report_rows
from scripts.feueron.models import Geschlecht, PersonRecord, PvBank, PvDb, generate_csv

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# FeuerON Bankverbindungen report parser
# ---------------------------------------------------------------------------


class BankverbindungEntry(NamedTuple):
    """A single person's bank details from a FeuerON Bankverbindungen report."""

    position: int
    nachname: str
    vorname: str
    personal_nr: str
    inhaber: str
    iban: str
    bic: str
    mandatsreferenz: str


def parse_bankverbindungen(path: Path) -> list[BankverbindungEntry]:
    """Parse a FeuerON Bankverbindungen CSV report.

    Each person spans 3 rows:
      1. Position, Name, Personal-Nr, Inhaber
      2. Bank, Ort, IBAN
      3. Mandatsreferenz, Erteilt am, BIC
    """
    entries: list[BankverbindungEntry] = []
    rows = list(iter_report_rows(path))

    i = 0
    while i < len(rows):
        row1, is_page1 = rows[i]
        first = row1[0].strip()

        if not first.isdigit():
            i += 1
            continue

        pos = int(first)

        # Row 1: [pos, '', name, ('',)? pnr, '', '', inhaber, '']
        if is_page1:
            name = row1[2].strip()
            personal_nr = row1[4].strip()
            inhaber = row1[7].strip()
        else:
            name = row1[2].strip()
            personal_nr = row1[3].strip()
            inhaber = row1[6].strip()

        # Row 2: bank details — ['', '', bank, ('',)? ort, '', '', iban, '']
        iban = ""
        if i + 1 < len(rows):
            row2, is_p1_r2 = rows[i + 1]
            if is_p1_r2:
                iban = row2[7].strip() if len(row2) > 7 else ""
            else:
                iban = row2[6].strip() if len(row2) > 6 else ""

        # Row 3: mandate — ['', '', mandatsref, ('',)? erteilt_am, '', '', bic, '']
        bic = ""
        mandatsreferenz = ""
        if i + 2 < len(rows):
            row3, is_p1_r3 = rows[i + 2]
            if is_p1_r3:
                mandatsreferenz = row3[2].strip() if len(row3) > 2 else ""
                bic = row3[7].strip() if len(row3) > 7 else ""
            else:
                mandatsreferenz = row3[2].strip() if len(row3) > 2 else ""
                bic = row3[6].strip() if len(row3) > 6 else ""

        # Split "Nachname, Vorname"
        if ", " in name:
            nachname, vorname = name.split(", ", 1)
        else:
            nachname = name
            vorname = ""

        entries.append(
            BankverbindungEntry(
                position=pos,
                nachname=nachname.strip(),
                vorname=vorname.strip(),
                personal_nr=personal_nr,
                inhaber=inhaber,
                iban=iban,
                bic=bic,
                mandatsreferenz=mandatsreferenz,
            )
        )

        i += 3
        continue

    return entries


# ---------------------------------------------------------------------------
# Fox112 Excel reader
# ---------------------------------------------------------------------------


class _Fox112BankPerson(NamedTuple):
    nachname: str
    vorname: str
    geschlecht: str
    geburtstag: str
    kontoinhaber: str
    iban: str
    bic: str
    mandatsreferenz: str


def _read_fox112_excel(path: Path) -> list[_Fox112BankPerson]:
    """Read bank-relevant columns from a Fox112 Excel export."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers) if name}

    required = {
        "NACHNAME",
        "VORNAME",
        "GESCHLECHT",
        "GEBURTSTAG",
        "ABTEILUNG",
        "KONTOINHABER",
        "IBAN",
        "BIC",
        "MANDATSREFERENZ",
    }
    missing = required - col.keys()
    if missing:
        msg = f"Fox112 Excel is missing columns: {', '.join(sorted(missing))}"
        raise ValueError(msg)

    persons: list[_Fox112BankPerson] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nachname = str(row[col["NACHNAME"]] or "").strip()
        vorname = str(row[col["VORNAME"]] or "").strip()
        if not nachname:
            continue
        abteilung = str(row[col["ABTEILUNG"]] or "").strip()
        if abteilung == "Externe Kontakte":
            continue
        persons.append(
            _Fox112BankPerson(
                nachname=nachname,
                vorname=vorname,
                geschlecht=str(row[col["GESCHLECHT"]] or "").strip(),
                geburtstag=str(row[col["GEBURTSTAG"]] or "").strip(),
                kontoinhaber=str(row[col["KONTOINHABER"]] or "").strip(),
                iban=str(row[col["IBAN"]] or "").strip(),
                bic=str(row[col["BIC"]] or "").strip(),
                mandatsreferenz=str(row[col["MANDATSREFERENZ"]] or "").strip(),
            )
        )

    wb.close()
    return persons


# ---------------------------------------------------------------------------
# Core: build bank-update CSV
# ---------------------------------------------------------------------------

_GESCHLECHT_MAP = {
    "M": Geschlecht.MAENNLICH,
    "W": Geschlecht.WEIBLICH,
    "J": Geschlecht.JURISTISCH,
}


def build_bank_update(
    fox112_excel: Path,
    feueron_bankverbindungen: Path,
    output: Path | io.StringIO,
    *,
    organisation: str = "Bröckel, OF",
    yes: bool = False,
) -> None:
    """Build a FeuerON-compatible CSV to update bank details from Fox112.

    Matches persons by name (case-insensitive).  Fox112 is the master for
    bank data; the FeuerON Bankverbindungen report provides the ``PERS_NR``.

    Persons without a Fox112 match or with ambiguous (duplicate) matches
    are skipped with a log warning.
    """
    # 1. Parse inputs
    feueron_entries = parse_bankverbindungen(feueron_bankverbindungen)
    fox112_persons = _read_fox112_excel(fox112_excel)

    # 2. Build Fox112 name index
    fox_by_name: dict[tuple[str, str], list[_Fox112BankPerson]] = {}
    for person in fox112_persons:
        key = (person.nachname.lower(), person.vorname.lower())
        fox_by_name.setdefault(key, []).append(person)

    # 3. Match and build PersonRecords
    records: list[PersonRecord] = []
    matched = 0
    skipped_no_match = 0
    skipped_duplicate = 0
    skipped_declined = 0
    skipped_unchanged = 0

    for entry in feueron_entries:
        key = (entry.nachname.lower(), entry.vorname.lower())
        candidates = fox_by_name.get(key, [])

        if len(candidates) == 0:
            logger.warning(
                "No Fox112 match for %s, %s (FeuerON PNR: %s) — skipped",
                entry.nachname,
                entry.vorname,
                entry.personal_nr,
            )
            skipped_no_match += 1
            continue

        if len(candidates) > 1:
            logger.warning(
                "Multiple Fox112 matches (%d) for %s, %s (FeuerON PNR: %s) — skipped",
                len(candidates),
                entry.nachname,
                entry.vorname,
                entry.personal_nr,
            )
            skipped_duplicate += 1
            continue

        fox = candidates[0]
        geschlecht = _GESCHLECHT_MAP.get(fox.geschlecht.upper())
        if geschlecht is None:
            logger.warning(
                "Unknown GESCHLECHT '%s' for %s, %s — skipped",
                fox.geschlecht,
                entry.nachname,
                entry.vorname,
            )
            continue

        # Detect changes
        changes: list[tuple[str, str, str]] = []
        if fox.kontoinhaber != entry.inhaber:
            changes.append(("Inhaber", entry.inhaber, fox.kontoinhaber))
        if fox.iban != entry.iban:
            changes.append(("IBAN", entry.iban, fox.iban))
        if fox.bic != entry.bic:
            changes.append(("BIC", entry.bic, fox.bic))
        if fox.mandatsreferenz != entry.mandatsreferenz:
            changes.append(
                ("Mandatsreferenz", entry.mandatsreferenz, fox.mandatsreferenz)
            )

        # Skip persons with no changes
        if not changes:
            skipped_unchanged += 1
            continue

        # Log and confirm changes
        logger.info(
            "Updating %s, %s (%s):",
            fox.nachname,
            fox.vorname,
            entry.personal_nr,
        )
        for label, old, new in changes:
            if old:
                logger.info("  %s: %s → %s", label, old, new)
            else:
                logger.info("  %s: (empty) → %s", label, new)

        if not yes and not typer.confirm("  Include in CSV?", default=True):
            skipped_declined += 1
            continue

        records.append(
            PersonRecord(
                stammdaten=PvDb(
                    ORGANISATION=organisation,
                    NACHNAME=fox.nachname,
                    VORNAME=fox.vorname,
                    PERS_NR=entry.personal_nr,
                    GEBURT=fox.geburtstag,
                    GESCHLECHT=geschlecht,
                ),
                bank=PvBank(
                    INHABER=fox.kontoinhaber or None,
                    IBAN=fox.iban or None,
                    BIC=fox.bic or None,
                    MANDATSREFERENZ=fox.mandatsreferenz or None,
                ),
            )
        )
        matched += 1

    logger.info(
        "Bank update: %d included, %d declined, %d unchanged, %d skipped (no match), %d skipped (duplicate)",
        matched,
        skipped_declined,
        skipped_unchanged,
        skipped_no_match,
        skipped_duplicate,
    )

    # 4. Generate FeuerON CSV
    generate_csv(records, output)
