"""Build a FeuerON address-update CSV from Fox112 master data.

Compares Fox112 addresses (Excel export) with the FeuerON Adressenliste
(CSV report) and outputs a FeuerON-compatible CSV import file to correct
addresses where they differ.  Fox112 is the source of truth; the FeuerON
Adressenliste provides the PERS_NR required for the import.
"""

from __future__ import annotations

import csv
import io
import logging
from pathlib import Path
from typing import NamedTuple

import openpyxl

from scripts.feueron.models import Geschlecht, PersonRecord, PvDb, generate_csv

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# FeuerON Adressenliste report parser
# ---------------------------------------------------------------------------


class AdressenlisteEntry(NamedTuple):
    """A single person row from a FeuerON Adressenliste report."""

    position: int
    nachname: str
    vorname: str
    personal_nr: str
    adresse: str


def parse_adressenliste(path: Path) -> list[AdressenlisteEntry]:
    """Parse a FeuerON Adressenliste CSV report.

    Handles the paginated export format with metadata headers, page footers
    (``Gesamtsumme``), and page headers (``Pos.``).  The column layout shifts
    between page 1 (9 columns) and subsequent pages (8 columns).
    """
    with open(path, encoding="utf-8-sig") as f:
        content = f.read()

    entries: list[AdressenlisteEntry] = []
    reader = csv.reader(io.StringIO(content), delimiter=";")

    for row in reader:
        if not row:
            continue

        # Data rows start with a position number
        first = row[0].strip()
        if not first.isdigit():
            continue

        pos = int(first)

        # Page 1 has 9 columns: [pos, '', name, '', pnr, '', addr, '', '']
        # Page 2+ has 8 columns: [pos, '', name, pnr, '', addr, '', '']
        if len(row) >= 9 and row[3] == "" and "_" in row[4]:
            name = row[2].strip()
            personal_nr = row[4].strip()
            adresse = row[6].strip()
        else:
            name = row[2].strip()
            personal_nr = row[3].strip()
            adresse = row[5].strip()

        # Split "Nachname, Vorname" into components
        if ", " in name:
            nachname, vorname = name.split(", ", 1)
        else:
            nachname = name
            vorname = ""

        entries.append(
            AdressenlisteEntry(
                position=pos,
                nachname=nachname.strip(),
                vorname=vorname.strip(),
                personal_nr=personal_nr,
                adresse=adresse,
            )
        )

    return entries


# ---------------------------------------------------------------------------
# Fox112 Excel reader
# ---------------------------------------------------------------------------


class _Fox112Person(NamedTuple):
    nachname: str
    vorname: str
    strasse: str
    hausnr: str
    plz: str
    ort: str
    geschlecht: str
    geburtstag: str


def _read_fox112_excel(path: Path) -> list[_Fox112Person]:
    """Read address-relevant columns from a Fox112 Excel export."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active

    # Build header index
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    col = {name: idx for idx, name in enumerate(headers) if name}

    required = {
        "NACHNAME",
        "VORNAME",
        "STRASSE",
        "HAUSNR",
        "PLZ",
        "ORT",
        "GESCHLECHT",
        "GEBURTSTAG",
    }
    missing = required - col.keys()
    if missing:
        msg = f"Fox112 Excel is missing columns: {', '.join(sorted(missing))}"
        raise ValueError(msg)

    persons: list[_Fox112Person] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nachname = str(row[col["NACHNAME"]] or "").strip()
        vorname = str(row[col["VORNAME"]] or "").strip()
        if not nachname:
            continue
        persons.append(
            _Fox112Person(
                nachname=nachname,
                vorname=vorname,
                strasse=str(row[col["STRASSE"]] or "").strip(),
                hausnr=str(row[col["HAUSNR"]] or "").strip(),
                plz=str(row[col["PLZ"]] or "").strip(),
                ort=str(row[col["ORT"]] or "").strip(),
                geschlecht=str(row[col["GESCHLECHT"]] or "").strip(),
                geburtstag=str(row[col["GEBURTSTAG"]] or "").strip(),
            )
        )

    wb.close()
    return persons


# ---------------------------------------------------------------------------
# Core: build address-update CSV
# ---------------------------------------------------------------------------

_GESCHLECHT_MAP = {
    "M": Geschlecht.MAENNLICH,
    "W": Geschlecht.WEIBLICH,
    "J": Geschlecht.JURISTISCH,
}


def build_address_update(
    fox112_excel: Path,
    feueron_adressenliste: Path,
    output: Path | io.StringIO,
    *,
    organisation: str = "Bröckel, OF",
) -> None:
    """Build a FeuerON-compatible CSV to update addresses from Fox112.

    Matches persons by name (case-insensitive).  Fox112 is the master for
    address data; the FeuerON Adressenliste provides the ``PERS_NR``.

    Persons without a Fox112 match or with ambiguous (duplicate) matches
    are skipped with a log warning.
    """
    # 1. Parse inputs
    feueron_entries = parse_adressenliste(feueron_adressenliste)
    fox112_persons = _read_fox112_excel(fox112_excel)

    # 2. Build Fox112 name index: (nachname_lower, vorname_lower) -> list[_Fox112Person]
    fox_by_name: dict[tuple[str, str], list[_Fox112Person]] = {}
    for person in fox112_persons:
        key = (person.nachname.lower(), person.vorname.lower())
        fox_by_name.setdefault(key, []).append(person)

    # 3. Match and build PersonRecords
    records: list[PersonRecord] = []
    matched = 0
    skipped_no_match = 0
    skipped_duplicate = 0

    for entry in feueron_entries:
        key = (entry.nachname.lower(), entry.vorname.lower())
        candidates = fox_by_name.get(key, [])

        if len(candidates) == 0:
            logger.warning(
                "No Fox112 match for %s, %s (FeuerON PNR: %s) — skipped",
                entry.nachname,
                entry.vorname,
                entry.personal_nr,
            )
            skipped_no_match += 1
            continue

        if len(candidates) > 1:
            logger.warning(
                "Multiple Fox112 matches (%d) for %s, %s (FeuerON PNR: %s) — skipped",
                len(candidates),
                entry.nachname,
                entry.vorname,
                entry.personal_nr,
            )
            skipped_duplicate += 1
            continue

        fox = candidates[0]
        geschlecht = _GESCHLECHT_MAP.get(fox.geschlecht.upper())
        if geschlecht is None:
            logger.warning(
                "Unknown GESCHLECHT '%s' for %s, %s — skipped",
                fox.geschlecht,
                entry.nachname,
                entry.vorname,
            )
            continue

        fox_address = f"{fox.strasse} {fox.hausnr}, {fox.plz} {fox.ort}".strip()
        if fox_address != entry.adresse:
            logger.info(
                "Updating %s, %s (%s): %s → %s",
                fox.nachname,
                fox.vorname,
                entry.personal_nr,
                entry.adresse,
                fox_address,
            )

        records.append(
            PersonRecord(
                stammdaten=PvDb(
                    ORGANISATION=organisation,
                    NACHNAME=fox.nachname,
                    VORNAME=fox.vorname,
                    PERS_NR=entry.personal_nr,
                    GEBURT=fox.geburtstag,
                    GESCHLECHT=geschlecht,
                    STRASSE=fox.strasse or None,
                    HAUSNR=fox.hausnr or None,
                    PLZ=fox.plz or None,
                    ORT=fox.ort or None,
                ),
            )
        )
        matched += 1

    logger.info(
        "Address update: %d matched, %d skipped (no match), %d skipped (duplicate)",
        matched,
        skipped_no_match,
        skipped_duplicate,
    )

    # 4. Generate FeuerON CSV
    generate_csv(records, output)
